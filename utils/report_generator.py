import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class TestResult:
    def __init__(self, test_id, test_name, status, error_message="",
                 screenshot_path="", execution_time=0.0, browser_name="chromium"):
        self.test_id = test_id
        self.test_name = test_name
        self.status = status
        self.error_message = error_message
        self.screenshot_path = screenshot_path
        self.execution_time = execution_time
        self.browser_name = browser_name


class ExcelReportGenerator:
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    PASS_FONT = Font(name="Calibri", size=10, bold=True, color="006100")
    PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FAIL_FONT = Font(name="Calibri", size=10, bold=True, color="9C0006")
    FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    SUBTITLE_FONT = Font(name="Calibri", size=11, bold=True, color="2E75B6")
    DATA_FONT = Font(name="Calibri", size=10)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def __init__(self):
        self.results = []
        self.start_time = None
        self.end_time = None
        self.report_dir = "reports"

    def set_start_time(self):
        self.start_time = datetime.now()

    def set_end_time(self):
        self.end_time = datetime.now()

    def add_result(self, result):
        self.results.append(result)

    def _get_report_filename(self):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"Preliminary_Form_Test_Report_{timestamp}.xlsx"

    def _calculate_statistics(self):
        total = len(self.results)
        passed = sum(1 for r in self.results if r.status == "PASS")
        failed = total - passed
        pass_pct = (passed / total * 100) if total > 0 else 0
        fail_pct = (failed / total * 100) if total > 0 else 0
        if self.start_time and self.end_time:
            duration = self.end_time - self.start_time
            duration_str = str(duration).split(".")[0]
        else:
            duration_str = "N/A"
        return {
            "total": total, "passed": passed, "failed": failed,
            "pass_percentage": round(pass_pct, 2),
            "fail_percentage": round(fail_pct, 2),
            "execution_date": self.start_time.strftime("%Y-%m-%d") if self.start_time else "N/A",
            "execution_time": self.start_time.strftime("%H:%M:%S") if self.start_time else "N/A",
            "duration": duration_str,
        }

    def _auto_adjust_column_width(self, worksheet):
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass
            adjusted_width = min(max_length + 3, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _create_summary_sheet(self, workbook):
        ws = workbook.active
        ws.title = "Summary"
        stats = self._calculate_statistics()
        ws.merge_cells("A1:D1")
        ws["A1"] = "CISCE Preliminary Affiliation Form - Test Execution Report"
        ws["A1"].font = self.TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:D2")
        ws["A2"] = "Automated Test Execution Summary"
        ws["A2"].font = self.SUBTITLE_FONT
        ws["A2"].alignment = Alignment(horizontal="center")
        current_row = 4
        summary_data = [
            ("Execution Date", stats["execution_date"]),
            ("Execution Time", stats["execution_time"]),
            ("Total Duration", stats["duration"]),
            ("", ""),
            ("Total Test Cases", stats["total"]),
            ("Passed Test Cases", stats["passed"]),
            ("Failed Test Cases", stats["failed"]),
            ("Pass Percentage", f"{stats['pass_percentage']}%"),
            ("Fail Percentage", f"{stats['fail_percentage']}%"),
        ]
        for label, value in summary_data:
            ws.cell(row=current_row, column=2, value=label).font = Font(name="Calibri", size=11, bold=True)
            ws.cell(row=current_row, column=2).border = self.THIN_BORDER
            value_cell = ws.cell(row=current_row, column=3, value=value)
            value_cell.font = self.DATA_FONT
            value_cell.border = self.THIN_BORDER
            if label == "Passed Test Cases":
                value_cell.font = self.PASS_FONT
                value_cell.fill = self.PASS_FILL
            elif label == "Failed Test Cases":
                value_cell.font = self.FAIL_FONT
                value_cell.fill = self.FAIL_FILL
            elif label == "Pass Percentage":
                value_cell.font = self.PASS_FONT
            elif label == "Fail Percentage":
                value_cell.font = self.FAIL_FONT
            current_row += 1
        current_row += 2
        ws.cell(row=current_row, column=2,
                value="Report generated automatically by pytest hooks").font = Font(
            name="Calibri", size=9, italic=True, color="808080")
        self._auto_adjust_column_width(ws)

    def _create_detailed_sheet(self, workbook):
        ws = workbook.create_sheet("Detailed Results")
        headers = ["Test Case ID", "Test Name", "Status", "Error Message",
                   "Screenshot Path", "Execution Time (s)", "Browser Name"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER
        for row_idx, result in enumerate(self.results, start=2):
            ws.cell(row=row_idx, column=1, value=result.test_id).font = self.DATA_FONT
            ws.cell(row=row_idx, column=1).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=2, value=result.test_name).font = self.DATA_FONT
            ws.cell(row=row_idx, column=2).border = self.THIN_BORDER
            status_cell = ws.cell(row=row_idx, column=3, value=result.status)
            status_cell.alignment = Alignment(horizontal="center")
            status_cell.border = self.THIN_BORDER
            if result.status == "PASS":
                status_cell.font = self.PASS_FONT
                status_cell.fill = self.PASS_FILL
            else:
                status_cell.font = self.FAIL_FONT
                status_cell.fill = self.FAIL_FILL
            error_msg = result.error_message[:500] if result.error_message else ""
            ws.cell(row=row_idx, column=4, value=error_msg).font = self.DATA_FONT
            ws.cell(row=row_idx, column=4).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=5, value=result.screenshot_path).font = self.DATA_FONT
            ws.cell(row=row_idx, column=5).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=6, value=round(result.execution_time, 2)).font = self.DATA_FONT
            ws.cell(row=row_idx, column=6).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=7, value=result.browser_name).font = self.DATA_FONT
            ws.cell(row=row_idx, column=7).border = self.THIN_BORDER
        ws.freeze_panes = "A2"
        self._auto_adjust_column_width(ws)

    def generate_report(self):
        os.makedirs(self.report_dir, exist_ok=True)
        workbook = Workbook()
        self._create_summary_sheet(workbook)
        self._create_detailed_sheet(workbook)
        filename = self._get_report_filename()
        report_path = os.path.join(self.report_dir, filename)
        workbook.save(report_path)
        print(f"\n{'='*60}")
        print(f"  EXCEL TEST REPORT GENERATED SUCCESSFULLY")
        print(f"  Location: {os.path.abspath(report_path)}")
        print(f"{'='*60}\n")
        return report_path
